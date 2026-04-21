"""
ASCENT — Flask version for Azure deployment
WRI India | Full-sector GHG emissions questionnaire with IPCC 2019 formulas
"""
from flask import Flask, render_template, request, jsonify, send_file
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.utils
import json
import io
import os

app = Flask(__name__)

# ─── EMISSION FACTORS (IPCC 2019 / CEA 2023) ──────────────────────────────────
FUEL_EF = {
    "Electricity":        (0.82,   0.0,     0.0,     None,    None),
    "Coal":               (94.6,   0.0003,  0.0015,  0.02326, 1.3),
    "Firewood":           (112.0,  0.03,    0.004,   0.01560, 0.65),
    "Kerosene":           (71.9,   0.003,   0.0006,  0.04360, 0.80),
    "LPG":                (63.1,   0.0003,  0.0001,  0.04740, 0.55),
    "PNG":                (56.1,   0.001,   0.0001,  0.04800, None),
    "Petrol":             (69.3,   0.003,   0.0006,  0.04470, 0.742),
    "Diesel":             (74.1,   0.003,   0.0006,  0.04260, 0.840),
    "CNG":                (56.1,   0.001,   0.0001,  0.04740, None),
    "Auto LPG":           (63.1,   0.0003,  0.0001,  0.04740, 0.55),
    "Aviation gasoline":  (70.0,   0.003,   0.0006,  0.04400, 0.72),
    "Jet kerosene":       (71.9,   0.003,   0.0006,  0.04360, 0.80),
    "Natural gas (TJ)":   (56.1,   0.001,   0.0001,  1.0,     None),
    "MSW incineration":   (0.69,   0.000013,0.000032,None,    None),
}
GWP = {"CO2": 1.0, "CH4": 28.0, "N2O": 265.0}

SECTOR_COLORS = {
    "Buildings – Residential":    "#C0392B",
    "Buildings – Commercial":     "#E74C3C",
    "Buildings – Public & Inst.": "#F1948A",
    "Buildings – Industrial":     "#FADBD8",
    "Electricity Generation":     "#922B21",
    "Transport – Road":           "#E67E22",
    "Transport – Rail":           "#F39C12",
    "Transport – Water/Aviation": "#F9E4B7",
    "Waste – Solid Waste":        "#16A085",
    "Waste – Biological":         "#1ABC9C",
    "Waste – Wastewater":         "#4AADA8",
    "AFOLU":                      "#27AE60",
    "IPPU":                       "#8E44AD",
}
UNIT_COST = {
    "Buildings – Residential":    10.0,
    "Buildings – Commercial":     11.0,
    "Buildings – Public & Inst.":  9.0,
    "Buildings – Industrial":     14.0,
    "Electricity Generation":     8.0,
    "Transport – Road":           18.0,
    "Transport – Rail":           15.0,
    "Transport – Water/Aviation": 12.0,
    "Waste – Solid Waste":         8.0,
    "Waste – Biological":          6.0,
    "Waste – Wastewater":          9.5,
    "AFOLU":                        3.5,
    "IPPU":                        22.0,
}
SECTORS = list(SECTOR_COLORS.keys())

INDIA_CITIES = [
    {"state":"Andaman & Nicobar Islands","district":"South Andaman","city":"Port Blair","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anakapalli","city":"Narsipatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anakapalli","city":"Yelamanchili","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Anantapur","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Dharmavaram","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Gooty","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Guntakal","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Kadiri","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Kalyandurgam","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Rayadurg","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Anantapur","city":"Tadipatri","climate":"Hot and Dry"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"B Kothakota","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Madanapalle","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Rajampeta","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Annamaiya","city":"Rayachoty","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Bapatla","city":"Addanki","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Bapatla","city":"Repalle","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Chittoor","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Kuppam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Nagari","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Palamaner","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Chittoor","city":"Punganur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Kovvur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Nidadavole","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"East Godavari","city":"Rajahmundry","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Chintalapudi","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Eluru","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Jangareddygudem","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Eluru","city":"Nuzivid","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Bapatla","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Guntur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Mangalagiri Tadepalli","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Ponnur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Guntur","city":"Tenali","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Badvel","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Jammalamadugu","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Kadapa","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Kamalapuram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Mydukur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Proddatur","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kadapa","city":"Pulivendula","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Kakinada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Pithapuram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kakinada","city":"Samalkot","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Gudivada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Machilipatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Krishna","city":"Vijayawada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kurnool","city":"Adoni","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Kurnool","city":"Kurnool","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"NTR","city":"Vijayawada","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Nellore","city":"Nellore","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Prakasam","city":"Ongole","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Srikakulam","city":"Srikakulam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Tirupati","city":"Tirupati","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Visakhapatnam","city":"Gvmc Visakhapatnam","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"Vizianagaram","city":"Vizianagaram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"West Godavari","city":"Bhimavaram","climate":"Warm & humid"},
    {"state":"Andhra Pradesh","district":"West Godavari","city":"Tadepalligudem","climate":"Warm & humid"},
    {"state":"Arunachal Pradesh","district":"East Siang","city":"Pasighat","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"Papum Pare","city":"Itanagar","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"Tawang","city":"Tawang","climate":"Cold"},
    {"state":"Arunachal Pradesh","district":"West Kameng","city":"Bomdila","climate":"Cold"},
    {"state":"Assam","district":"Barpeta","city":"Barpeta","climate":"Warm & humid"},
    {"state":"Assam","district":"Barpeta","city":"Barpeta Road","climate":"Warm & humid"},
    {"state":"Assam","district":"Bongaigaon","city":"Bongaigaon","climate":"Warm & humid"},
    {"state":"Assam","district":"Cachar","city":"Silchar","climate":"Warm & humid"},
    {"state":"Assam","district":"Dibrugarh","city":"Dibrugarh","climate":"Warm & humid"},
    {"state":"Assam","district":"Golaghat","city":"Golaghat","climate":"Warm & humid"},
    {"state":"Assam","district":"Jorhat","city":"Jorhat","climate":"Warm & humid"},
    {"state":"Assam","district":"Kamrup","city":"Guwahati","climate":"Warm & humid"},
    {"state":"Assam","district":"Kamrup","city":"North Guwahati","climate":"Warm & humid"},
    {"state":"Assam","district":"Karbi Anlong","city":"Diphu","climate":"Warm & humid"},
    {"state":"Assam","district":"Karimganj","city":"Karimganj","climate":"Warm & humid"},
    {"state":"Assam","district":"Kokrajhar","city":"Kokrajhar","climate":"Warm & humid"},
    {"state":"Assam","district":"Nagaon","city":"Nagaon","climate":"Warm & humid"},
    {"state":"Assam","district":"Nalbari","city":"Nalbari","climate":"Warm & humid"},
    {"state":"Assam","district":"North Lakhimpur","city":"North Lakhimpur","climate":"Warm & humid"},
    {"state":"Assam","district":"Sibsagar","city":"Sibsagar","climate":"Warm & humid"},
    {"state":"Assam","district":"Sonitpur","city":"Tezpur","climate":"Warm & humid"},
    {"state":"Assam","district":"Tinsukia","city":"Tinsukia","climate":"Warm & humid"},
    {"state":"Bihar","district":"Araria","city":"Araria","climate":"Composite"},
    {"state":"Bihar","district":"Aurangabad","city":"Aurangabad","climate":"Composite"},
    {"state":"Bihar","district":"Begusarai","city":"Begusarai","climate":"Composite"},
    {"state":"Bihar","district":"Bhagalpur","city":"Bhagalpur","climate":"Composite"},
    {"state":"Bihar","district":"Bhojpur","city":"Ara","climate":"Composite"},
    {"state":"Bihar","district":"Buxar","city":"Buxar","climate":"Composite"},
    {"state":"Bihar","district":"Darbhanga","city":"Darbhanga","climate":"Composite"},
    {"state":"Bihar","district":"Gaya","city":"Bodh Gaya","climate":"Composite"},
    {"state":"Bihar","district":"Gaya","city":"Gaya","climate":"Composite"},
    {"state":"Bihar","district":"Gopalganj","city":"Gopalganj","climate":"Composite"},
    {"state":"Bihar","district":"Jahanabad","city":"Jehanabad","climate":"Composite"},
    {"state":"Bihar","district":"Jamui","city":"Jamui","climate":"Composite"},
    {"state":"Bihar","district":"Katihar","city":"Katihar","climate":"Composite"},
    {"state":"Bihar","district":"Khagaria","city":"Khagaria","climate":"Composite"},
    {"state":"Bihar","district":"Kishanganj","city":"Kishanganj","climate":"Composite"},
    {"state":"Bihar","district":"Lakhisarai","city":"Lakhisarai","climate":"Composite"},
    {"state":"Bihar","district":"Madhepura","city":"Madhepura","climate":"Composite"},
    {"state":"Bihar","district":"Madhubani","city":"Madhubani","climate":"Composite"},
    {"state":"Bihar","district":"Munger","city":"Munger","climate":"Composite"},
    {"state":"Bihar","district":"Muzaffarpur","city":"Muzaffarpur","climate":"Composite"},
    {"state":"Bihar","district":"Nalanda","city":"Biharsharif","climate":"Composite"},
    {"state":"Bihar","district":"Nalanda","city":"Rajgir","climate":"Composite"},
    {"state":"Bihar","district":"Nawada","city":"Nawada","climate":"Composite"},
    {"state":"Bihar","district":"Paschim Champaran","city":"Bettiah","climate":"Composite"},
    {"state":"Bihar","district":"Patna","city":"Patna","climate":"Composite"},
    {"state":"Bihar","district":"Purnia","city":"Purnia","climate":"Composite"},
    {"state":"Bihar","district":"Purvi Champaran","city":"Motihari","climate":"Composite"},
    {"state":"Bihar","district":"Rohtas","city":"Sasaram","climate":"Composite"},
    {"state":"Bihar","district":"Saharsa","city":"Saharsa","climate":"Composite"},
    {"state":"Bihar","district":"Samastipur","city":"Samastipur","climate":"Composite"},
    {"state":"Bihar","district":"Saran","city":"Chapra","climate":"Composite"},
    {"state":"Bihar","district":"Sheohar","city":"Sheohar","climate":"Composite"},
    {"state":"Bihar","district":"Sitamarhi","city":"Sitamarhi","climate":"Composite"},
    {"state":"Bihar","district":"Siwan","city":"Siwan","climate":"Composite"},
    {"state":"Bihar","district":"Supaul","city":"Supaul","climate":"Composite"},
    {"state":"Bihar","district":"Vaishali","city":"Hajipur","climate":"Composite"},
    {"state":"Chandigarh","district":"Chandigarh","city":"Chandigarh","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Balod","city":"Balod","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Bilaspur","city":"Bilaspur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Durg","city":"Bhilai Nagar","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Durg","city":"Durg","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Jagdalpur","city":"Jagdalpur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Janjgir Champa","city":"Champa","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Korba","city":"Korba","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Raigarh","city":"Raigarh","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Raipur","city":"Raipur","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Rajnandgaon","city":"Rajnandgaon","climate":"Composite"},
    {"state":"Chhattisgarh","district":"Sarguja","city":"Ambikapur","climate":"Composite"},
    {"state":"Dadra & Nagar Haveli","district":"Dadra And Nagar Haveli","city":"Silvassa","climate":"Warm & humid"},
    {"state":"Daman & Diu","district":"Daman","city":"Daman","climate":"Warm & humid"},
    {"state":"Daman & Diu","district":"Diu","city":"Diu","climate":"Warm & humid"},
    {"state":"Delhi","district":"New Delhi","city":"New Delhi","climate":"Composite"},
    {"state":"Delhi","district":"South Delhi","city":"Municipal Corporation Of Delhi","climate":"Composite"},
    {"state":"Goa","district":"North Goa","city":"Mapusa","climate":"Warm & humid"},
    {"state":"Goa","district":"North Goa","city":"Panaji","climate":"Warm & humid"},
    {"state":"Goa","district":"South Goa","city":"Margao","climate":"Warm & humid"},
    {"state":"Goa","district":"South Goa","city":"Mormugao","climate":"Warm & humid"},
    {"state":"Gujarat","district":"Ahmedabad","city":"Ahmedabad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Amreli","city":"Amreli","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Anand","city":"Anand","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Anand","city":"Vallabh Vidhyanagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Aravalli","city":"Modasa","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Banas Kantha","city":"Palanpur","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bharuch","city":"Ankleshwer","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bharuch","city":"Bharuch","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Bhavnagar","city":"Bhavnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Botad","city":"Botad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Chhota Udepure","city":"Chhota Udepur","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Dahod","city":"Dahod","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Devbhoomi Dwarka","city":"Dwarka","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gandhinagar","city":"Gandhinagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gandhinagar","city":"Kalol","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Gir Somnath","city":"Veraval","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Jamnagar","city":"Jamnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Junagadh","city":"Junagadh","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kachchh","city":"Bhuj","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kachchh","city":"Gandhidham","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Kheda","city":"Nadiad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Mahesana","city":"Mahesana","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Morbi","city":"Morbi","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Narmada","city":"Rajpipla","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Navsari","city":"Navsari","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Panch Mahals","city":"Godhra","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Patan","city":"Patan","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Porbandar","city":"Porbandar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Rajkot","city":"Gondal","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Rajkot","city":"Rajkot","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Sabar Kantha","city":"Himmatnagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surat","city":"Bardoli","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surat","city":"Surat","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Surendranagar Dudhrej","city":"Surendranagar","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Vadodara","city":"Vadodara","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Valsad","city":"Valsad","climate":"Hot and Dry"},
    {"state":"Gujarat","district":"Valsad","city":"Vapi","climate":"Hot and Dry"},
    {"state":"Haryana","district":"Ambala","city":"Ambala","climate":"Composite"},
    {"state":"Haryana","district":"Bhiwani","city":"Bhiwani","climate":"Composite"},
    {"state":"Haryana","district":"Faridabad","city":"Faridabad","climate":"Composite"},
    {"state":"Haryana","district":"Gurugram","city":"Gurugram","climate":"Composite"},
    {"state":"Haryana","district":"Hisar","city":"Hisar","climate":"Composite"},
    {"state":"Haryana","district":"Jhajjar","city":"Bahadurgarh","climate":"Composite"},
    {"state":"Haryana","district":"Jind","city":"Jind","climate":"Composite"},
    {"state":"Haryana","district":"Kaithal","city":"Kaithal","climate":"Composite"},
    {"state":"Haryana","district":"Karnal","city":"Karnal","climate":"Composite"},
    {"state":"Haryana","district":"Kurukshetra","city":"Thanesar","climate":"Composite"},
    {"state":"Haryana","district":"Mahendragarh","city":"Narnaul","climate":"Composite"},
    {"state":"Haryana","district":"Nuh","city":"Nuh","climate":"Composite"},
    {"state":"Haryana","district":"Palwal","city":"Palwal","climate":"Composite"},
    {"state":"Haryana","district":"Panchkula","city":"Panchkula","climate":"Composite"},
    {"state":"Haryana","district":"Panipat","city":"Panipat","climate":"Composite"},
    {"state":"Haryana","district":"Rewari","city":"Rewari","climate":"Composite"},
    {"state":"Haryana","district":"Rohtak","city":"Rohtak","climate":"Composite"},
    {"state":"Haryana","district":"Sirsa","city":"Sirsa","climate":"Composite"},
    {"state":"Haryana","district":"Sonipat","city":"Sonipat","climate":"Composite"},
    {"state":"Haryana","district":"Yamunanagar","city":"Yamunanagar","climate":"Composite"},
    {"state":"Himachal Pradesh","district":"Bilaspur","city":"Bilaspur","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Chamba","city":"Chamba","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Hamirpur","city":"Hamirpur","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Kangra","city":"Dharmsala","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Kangra","city":"Kangra","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Kullu","city":"Kullu","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Kullu","city":"Manali","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Mandi","city":"Mandi","climate":"Temperate"},
    {"state":"Himachal Pradesh","district":"Shimla","city":"Shimla","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Sirmour","city":"Nahan","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Solan","city":"Solan","climate":"Cold"},
    {"state":"Himachal Pradesh","district":"Una","city":"Una","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Anantnag","city":"Anantnag","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Badgam","city":"Badgam","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Baramula","city":"Baramula","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Baramula","city":"Sopore","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Doda","city":"Doda","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Jammu","city":"Jammu","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kathua","city":"Kathua","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kulgam","city":"Kulgam","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Kupwara","city":"Kupwara","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Pulwama","city":"Pulwama","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Punch","city":"Punch","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Rajouri","city":"Rajouri","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Srinagar","city":"Srinagar","climate":"Cold"},
    {"state":"Jammu & Kashmir","district":"Udhampur","city":"Udhampur","climate":"Cold"},
    {"state":"Jharkhand","district":"Bokaro","city":"Chas","climate":"Composite"},
    {"state":"Jharkhand","district":"Chatra","city":"Chatra","climate":"Composite"},
    {"state":"Jharkhand","district":"Deoghar","city":"Deoghar","climate":"Composite"},
    {"state":"Jharkhand","district":"Dhanbad","city":"Dhanbad","climate":"Composite"},
    {"state":"Jharkhand","district":"Dumka","city":"Dumka","climate":"Composite"},
    {"state":"Jharkhand","district":"East Singhbhum","city":"Jamshedpur","climate":"Composite"},
    {"state":"Jharkhand","district":"Garhwa","city":"Garhwa","climate":"Composite"},
    {"state":"Jharkhand","district":"Giridih","city":"Giridih","climate":"Composite"},
    {"state":"Jharkhand","district":"Godda","city":"Godda","climate":"Composite"},
    {"state":"Jharkhand","district":"Gumla","city":"Gumla","climate":"Composite"},
    {"state":"Jharkhand","district":"Hazaribagh","city":"Hazaribagh","climate":"Composite"},
    {"state":"Jharkhand","district":"Jamtara","city":"Jamtara","climate":"Composite"},
    {"state":"Jharkhand","district":"Koderma","city":"Koderma","climate":"Composite"},
    {"state":"Jharkhand","district":"Latehar","city":"Latehar","climate":"Composite"},
    {"state":"Jharkhand","district":"Lohardaga","city":"Lohardaga","climate":"Composite"},
    {"state":"Jharkhand","district":"Pakur","city":"Pakur","climate":"Composite"},
    {"state":"Jharkhand","district":"Palamu","city":"Medininagar","climate":"Composite"},
    {"state":"Jharkhand","district":"Ramgarh","city":"Ramgarh Nagar Parishad","climate":"Composite"},
    {"state":"Jharkhand","district":"Ranchi","city":"Ranchi","climate":"Composite"},
    {"state":"Jharkhand","district":"Sahebganj","city":"Sahibganj","climate":"Composite"},
    {"state":"Jharkhand","district":"Saraikela - Kharswan","city":"Seraikela","climate":"Composite"},
    {"state":"Jharkhand","district":"Simdega","city":"Simdega","climate":"Composite"},
    {"state":"Jharkhand","district":"West Singhbhum","city":"Chaibasa","climate":"Composite"},
    {"state":"Karnataka","district":"Bagalkote","city":"Bagalkot","climate":"Composite"},
    {"state":"Karnataka","district":"Ballary","city":"Bellary","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Belagavi","city":"Belgaum","climate":"Composite"},
    {"state":"Karnataka","district":"Bengaluru Rural","city":"Devanahalli","climate":"Composite"},
    {"state":"Karnataka","district":"Bengaluru Urban","city":"Bruhat Bengaluru Mahanagara Palike","climate":"Composite"},
    {"state":"Karnataka","district":"Bidar","city":"Bidar","climate":"Composite"},
    {"state":"Karnataka","district":"Chamarajanagara","city":"Chamarajanagar","climate":"Composite"},
    {"state":"Karnataka","district":"Chikkaballapura","city":"Chikkaballapura","climate":"Composite"},
    {"state":"Karnataka","district":"Chikkamagaluru","city":"Chikmagalur","climate":"Composite"},
    {"state":"Karnataka","district":"Chitradurga","city":"Chitradurga","climate":"Composite"},
    {"state":"Karnataka","district":"Dakshina kannada","city":"Mangalore","climate":"Composite"},
    {"state":"Karnataka","district":"Davangere","city":"Davanagere","climate":"Composite"},
    {"state":"Karnataka","district":"Dharwada","city":"Hubli-Dharwad","climate":"Composite"},
    {"state":"Karnataka","district":"Gadag","city":"Gadag-Betigeri","climate":"Composite"},
    {"state":"Karnataka","district":"Hassan","city":"Hassan","climate":"Composite"},
    {"state":"Karnataka","district":"Haveri","city":"Haveri","climate":"Composite"},
    {"state":"Karnataka","district":"Kalaburagi","city":"Gulbarga","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Kodagu","city":"Madikeri","climate":"Composite"},
    {"state":"Karnataka","district":"Kolar","city":"Kolar","climate":"Composite"},
    {"state":"Karnataka","district":"Koppal","city":"Koppal","climate":"Composite"},
    {"state":"Karnataka","district":"Mandya","city":"Mandya","climate":"Composite"},
    {"state":"Karnataka","district":"Mysuru","city":"Mysore","climate":"Composite"},
    {"state":"Karnataka","district":"Raichur","city":"Raichur","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Ramanagara","city":"Ramanagara","climate":"Composite"},
    {"state":"Karnataka","district":"Shivamogga","city":"Shimoga","climate":"Composite"},
    {"state":"Karnataka","district":"Tumakuru","city":"Tumkur","climate":"Composite"},
    {"state":"Karnataka","district":"Udupi","city":"Udupi","climate":"Composite"},
    {"state":"Karnataka","district":"Uttara Kannada","city":"Karwar","climate":"Composite"},
    {"state":"Karnataka","district":"Vijayapura","city":"Bijapur","climate":"Hot and Dry"},
    {"state":"Karnataka","district":"Yadgir","city":"Yadgir","climate":"Hot and Dry"},
    {"state":"Kerala","district":"Alappuzha","city":"Alappuzha","climate":"Warm & humid"},
    {"state":"Kerala","district":"Ernakulam","city":"Kochi","climate":"Warm & humid"},
    {"state":"Kerala","district":"Idukki","city":"Thodupuzha","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kannur","city":"Kannur","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kannur","city":"Thalassery","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kasaragod","city":"Kasaragod","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kollam","city":"Kollam","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kottayam","city":"Kottayam","climate":"Warm & humid"},
    {"state":"Kerala","district":"Kozhikode","city":"Kozhikode","climate":"Warm & humid"},
    {"state":"Kerala","district":"Malappuram","city":"Malappuram","climate":"Warm & humid"},
    {"state":"Kerala","district":"Palakkad","city":"Palakkad","climate":"Warm & humid"},
    {"state":"Kerala","district":"Pathanamthitta","city":"Pathanamthitta","climate":"Warm & humid"},
    {"state":"Kerala","district":"Thiruvananthapuram","city":"Thiruvananthapuram","climate":"Warm & humid"},
    {"state":"Kerala","district":"Thrissur","city":"Thrissur","climate":"Warm & humid"},
    {"state":"Kerala","district":"Wayanad","city":"Kalpetta","climate":"Warm & humid"},
    {"state":"Ladakh","district":"Kargil","city":"Kargil","climate":"Cold"},
    {"state":"Ladakh","district":"Leh","city":"Leh","climate":"Cold"},
    {"state":"Madhya Pradesh","district":"Agar Malwa","city":"Agar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Alirajpur","city":"Alirajpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Anuppur","city":"Anuppur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ashoknagar","city":"Ashoknagar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Balaghat","city":"Balaghat","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Barwani","city":"Badwani","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Betul","city":"Betul","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Bhind","city":"Bhind","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Bhopal","city":"Bhopal","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Burhanpur","city":"Burhanpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Chhatarpur","city":"Chhatarpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Chhindwara","city":"Chhindwara","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Damoh","city":"Damoh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Datia","city":"Datia","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Dewas","city":"Dewas","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Dhar","city":"Dhar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Guna","city":"Guna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Gwalior","city":"Gwalior","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Harda","city":"Harda","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Indore","city":"Indore","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Jabalpur","city":"Jabalpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Jhabua","city":"Jhabua","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Khandwa","city":"Khandwa","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Khargaon","city":"Khargone","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Mandla","city":"Mandla","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Mandsaur","city":"Mandsaur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Morena","city":"Morena","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Murwara (Katni)","city":"Katni","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Narmadapuram","city":"Narmadapuram","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Narsinghpur","city":"Narsinghpur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Neemuch","city":"Neemuch","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Panna","city":"Panna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Raisen","city":"Raisen","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Rajgarh","city":"Rajgarh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ratlam","city":"Ratlam","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Rewa","city":"Rewa","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sagar","city":"Sagar","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Satna","city":"Satna","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sehore","city":"Sehore","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Seoni","city":"Seoni","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shahdol","city":"Shahdol","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shajapur","city":"Shajapur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sheopur","city":"Sheopur","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Shivpuri","city":"Shivpuri","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Sidhi","city":"Sidhi","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Singrauli","city":"Singrauli","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Tikamgarh","city":"Tikamgarh","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Ujjain","city":"Ujjain","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Umaria","city":"Umaria","climate":"Composite"},
    {"state":"Madhya Pradesh","district":"Vidisha","city":"Vidisha","climate":"Composite"},
    {"state":"Maharashtra","district":"Ahmednagar","city":"Ahmednagar","climate":"Composite"},
    {"state":"Maharashtra","district":"Akola","city":"Akola","climate":"Composite"},
    {"state":"Maharashtra","district":"Amravati","city":"Amravati","climate":"Composite"},
    {"state":"Maharashtra","district":"Aurangabad","city":"Aurangabad","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Beed","city":"Beed","climate":"Composite"},
    {"state":"Maharashtra","district":"Bhandara","city":"Bhandara","climate":"Composite"},
    {"state":"Maharashtra","district":"Buldana","city":"Buldana","climate":"Composite"},
    {"state":"Maharashtra","district":"Chandrapur","city":"Chandrapur","climate":"Composite"},
    {"state":"Maharashtra","district":"Dhule","city":"Dhule","climate":"Composite"},
    {"state":"Maharashtra","district":"Gadchiroli","city":"Gadchiroli","climate":"Composite"},
    {"state":"Maharashtra","district":"Gondiya","city":"Gondiya","climate":"Composite"},
    {"state":"Maharashtra","district":"Hingoli","city":"Hingoli","climate":"Composite"},
    {"state":"Maharashtra","district":"Jalgaon","city":"Jalgaon","climate":"Composite"},
    {"state":"Maharashtra","district":"Jalna","city":"Jalna","climate":"Composite"},
    {"state":"Maharashtra","district":"Kolhapur","city":"Kolhapur","climate":"Composite"},
    {"state":"Maharashtra","district":"Latur","city":"Latur","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Mumbai","city":"Greater Mumbai","climate":"Composite"},
    {"state":"Maharashtra","district":"Nagpur","city":"Nagpur","climate":"Composite"},
    {"state":"Maharashtra","district":"Nanded Waghala","city":"Nanded Waghala","climate":"Composite"},
    {"state":"Maharashtra","district":"Nandurbar","city":"Nandurbar","climate":"Composite"},
    {"state":"Maharashtra","district":"Nashik","city":"Nashik","climate":"Composite"},
    {"state":"Maharashtra","district":"Osmanabad","city":"Osmanabad","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Palghar","city":"Palghar","climate":"Composite"},
    {"state":"Maharashtra","district":"Parbhani","city":"Parbhani","climate":"Composite"},
    {"state":"Maharashtra","district":"Pune","city":"Pune","climate":"Composite"},
    {"state":"Maharashtra","district":"Raigadh","city":"Alibag","climate":"Composite"},
    {"state":"Maharashtra","district":"Ratnagiri","city":"Ratnagiri","climate":"Composite"},
    {"state":"Maharashtra","district":"Sangli","city":"Sangli","climate":"Composite"},
    {"state":"Maharashtra","district":"Satara","city":"Satara","climate":"Composite"},
    {"state":"Maharashtra","district":"Sindhudurga","city":"Kudal","climate":"Composite"},
    {"state":"Maharashtra","district":"Solapur","city":"Solapur","climate":"Hot and Dry"},
    {"state":"Maharashtra","district":"Thane","city":"Thane","climate":"Composite"},
    {"state":"Maharashtra","district":"Wardha","city":"Wardha","climate":"Composite"},
    {"state":"Maharashtra","district":"Washim","city":"Washim","climate":"Composite"},
    {"state":"Maharashtra","district":"Yavatmal","city":"Yavatmal","climate":"Composite"},
    {"state":"Manipur","district":"Bishnupur","city":"Bishnupur","climate":"Warm & humid"},
    {"state":"Manipur","district":"Imphal East","city":"Imphal","climate":"Warm & humid"},
    {"state":"Manipur","district":"Thoubal","city":"Thoubal","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"East Garo Hills","city":"Williamnagar","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"East Khasi","city":"Shillong","climate":"Warm & humid"},
    {"state":"Meghalaya","district":"West Garo Hills","city":"Tura","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Aizawl","city":"Aizawl","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Champhai","city":"Champhai","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Kolasib","city":"Kolasib","climate":"Warm & humid"},
    {"state":"Mizoram","district":"Lunglei","city":"Lunglei","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Dimapur","city":"Dimapur","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Kohima","city":"Kohima","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Mokokchung","city":"Mokokchung","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Mon","city":"Mon","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Phek","city":"Phek","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Tuensang","city":"Tuensang","climate":"Warm & humid"},
    {"state":"Nagaland","district":"Wokha","city":"Wokha","climate":"Warm & humid"},
    {"state":"Odisha","district":"Anugul","city":"Anugul","climate":"Warm & humid"},
    {"state":"Odisha","district":"Balangir","city":"Balangir","climate":"Warm & humid"},
    {"state":"Odisha","district":"Baleshwar","city":"Baleshwar Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Bargarh","city":"Bargarh","climate":"Warm & humid"},
    {"state":"Odisha","district":"Bhadrak","city":"Bhadrak","climate":"Warm & humid"},
    {"state":"Odisha","district":"Cuttack","city":"Cuttack","climate":"Warm & humid"},
    {"state":"Odisha","district":"Dhenkanal","city":"Dhenkanal","climate":"Warm & humid"},
    {"state":"Odisha","district":"Gajapati","city":"Paralakhemundi","climate":"Warm & humid"},
    {"state":"Odisha","district":"Ganjam","city":"Brahmapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jagatsinghpur","city":"Jagatsinghapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jajpur","city":"Jajapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Jharsuguda","city":"Jharsuguda","climate":"Warm & humid"},
    {"state":"Odisha","district":"Kalahandi","city":"Bhawanipatna","climate":"Warm & humid"},
    {"state":"Odisha","district":"Kendrapara","city":"Kendrapara","climate":"Warm & humid"},
    {"state":"Odisha","district":"Keonjhar","city":"Kendujhar","climate":"Warm & humid"},
    {"state":"Odisha","district":"Khurda","city":"Bhubaneswar","climate":"Warm & humid"},
    {"state":"Odisha","district":"Koraput","city":"Koraput","climate":"Warm & humid"},
    {"state":"Odisha","district":"Malkangiri","city":"Malkangiri","climate":"Warm & humid"},
    {"state":"Odisha","district":"Mayurbhanj","city":"Baripada Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Nabarangapur","city":"Nabarangapur","climate":"Warm & humid"},
    {"state":"Odisha","district":"Nayagarh","city":"Nayagarh","climate":"Warm & humid"},
    {"state":"Odisha","district":"Puri","city":"Puri","climate":"Warm & humid"},
    {"state":"Odisha","district":"Rayagada","city":"Rayagada","climate":"Warm & humid"},
    {"state":"Odisha","district":"Sambalpur","city":"Sambalpur Town","climate":"Warm & humid"},
    {"state":"Odisha","district":"Sundergarh","city":"Raurkela Town","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Karaikal","city":"Karaikal","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Mahe","city":"Mahe","climate":"Warm & humid"},
    {"state":"Puducherry","district":"Puducherry","city":"Puducherry","climate":"Warm & humid"},
    {"state":"Punjab","district":"Amritsar","city":"Amritsar","climate":"Composite"},
    {"state":"Punjab","district":"Barnala","city":"Barnala","climate":"Composite"},
    {"state":"Punjab","district":"Bathinda","city":"Bathinda","climate":"Composite"},
    {"state":"Punjab","district":"Faridkot","city":"Faridkot","climate":"Composite"},
    {"state":"Punjab","district":"Fazilka","city":"Abohar","climate":"Composite"},
    {"state":"Punjab","district":"Firozpur","city":"Firozpur","climate":"Composite"},
    {"state":"Punjab","district":"Gurdaspur","city":"Gurdaspur","climate":"Composite"},
    {"state":"Punjab","district":"Hoshiarpur","city":"Hoshiarpur","climate":"Composite"},
    {"state":"Punjab","district":"Jalandhar","city":"Jalandhar","climate":"Composite"},
    {"state":"Punjab","district":"Kapurthala","city":"Kapurthala","climate":"Composite"},
    {"state":"Punjab","district":"Ludhiana","city":"Ludhiana","climate":"Composite"},
    {"state":"Punjab","district":"Mansa","city":"Mansa","climate":"Composite"},
    {"state":"Punjab","district":"Moga","city":"Moga","climate":"Composite"},
    {"state":"Punjab","district":"Muktsar","city":"Muktsar","climate":"Composite"},
    {"state":"Punjab","district":"Pathankot","city":"Pathankot","climate":"Composite"},
    {"state":"Punjab","district":"Patiala","city":"Patiala","climate":"Composite"},
    {"state":"Punjab","district":"Rupnagar","city":"Rupnagar","climate":"Composite"},
    {"state":"Punjab","district":"S.A.S. Nagar (Mohali)","city":"Mohali","climate":"Composite"},
    {"state":"Punjab","district":"Sangrur","city":"Sangrur","climate":"Composite"},
    {"state":"Punjab","district":"Sirhind Fatehgarh Sahib","city":"Sirhind Fatehgarh Sahib","climate":"Composite"},
    {"state":"Punjab","district":"Tarn Taran","city":"Tarn Taran","climate":"Composite"},
    {"state":"Rajasthan","district":"Ajmer","city":"Ajmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Alwar","city":"Alwar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Banswara","city":"Banswara","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Baran","city":"Baran","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Barmer","city":"Barmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bharatpur","city":"Bharatpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bhilwara","city":"Bhilwara","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bikaner","city":"Bikaner","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Bundi","city":"Bundi","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Chittaurgarh","city":"Chittaurgarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Churu","city":"Churu","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dausa","city":"Dausa","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dhaulpur","city":"Dhaulpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Dungarpur","city":"Dungarpur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Ganganagar","city":"Ganganagar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Hanumangarh","city":"Hanumangarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jaipur","city":"Jaipur Greater","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jaisalmer","city":"Jaisalmer","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jalor","city":"Jalor","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jhalawar","city":"Jhalawar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jhunjhunun","city":"Jhunjhunun","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Jodhpur","city":"Jodhpur North","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Karauli","city":"Karauli","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Kota","city":"Kota North","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Nagaur","city":"Nagaur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Pali","city":"Pali","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Pratapgarh","city":"Pratapgarh","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Rajsamand","city":"Rajsamand","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sawai Madhopur","city":"Sawai Madhopur","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sikar","city":"Sikar","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Sirohi","city":"Sirohi","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Tonk","city":"Tonk","climate":"Hot and Dry"},
    {"state":"Rajasthan","district":"Udaipur","city":"Udaipur","climate":"Hot and Dry"},
    {"state":"Sikkim","district":"Gangtok","city":"Gangtok","climate":"Cold"},
    {"state":"Sikkim","district":"Namchi","city":"Namchi","climate":"Cold"},
    {"state":"Tamil Nadu","district":"Ariyalur","city":"Ariyalur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Chengalpattu","city":"Chengalpattu","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Chennai","city":"Chennai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Coimbatore","city":"Coimbatore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Cuddalore","city":"Cuddalore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Dharmapuri","city":"Dharmapuri","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Dindigul","city":"Dindigul","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Erode","city":"Erode","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kallakurichi","city":"Kallakurichi","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kancheepuram","city":"Kancheepuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Kanniyakumari","city":"Nagarcoil","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Karur","city":"Karur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Krishnagiri","city":"Krishnagiri","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Madurai","city":"Madurai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Mayiladuthurai","city":"Mayiladuthurai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Nagapattinam","city":"Nagapattinam","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Namakkal","city":"Namakkal","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Perambalur","city":"Perambalur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Pudukottai","city":"Pudukottai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Ramanathapuram","city":"Ramanathapuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Ranipet","city":"Ranipet","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Salem","city":"Salem","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Sivaganga","city":"Sivagangai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tenkasi","city":"Tenkasi","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thanjavur","city":"Thanjavur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"The Nilgiris","city":"Udagamandalam","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Theni","city":"Theni Alinagaram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thirupathur","city":"Tirupathur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thiruvallur","city":"Tiruvallur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Thoothukkudi","city":"Thoothukudi Corporation","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tirunelveli","city":"Tirunelveli","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruppurur","city":"Tirupppur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruvannamalai","city":"Tiruvannamalai","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Tiruvarur","city":"Thiruvarur","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Trichy","city":"Tiruchirapalli","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Vellore","city":"Vellore","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Viluppuram","city":"Villupuram","climate":"Warm & humid"},
    {"state":"Tamil Nadu","district":"Virudhunagar","city":"Virudhunagar","climate":"Warm & humid"},
    {"state":"Telangana","district":"Adilabad","city":"Adilabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Bhadradri Kothagudem","city":"Kothagudem","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Hanumakonda","city":"Warangal","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Hyderabad","city":"Greater Hyderabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Jagitial","city":"Jagitial","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Jangaon","city":"Jangaon","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Karimnagar","city":"Karimnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Khammam","city":"Khammam","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mahabubabad","city":"Mahaboobabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mahabubnagar","city":"Mahabubnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Mancherial","city":"Mancherial","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Medak","city":"Medak","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nalgonda","city":"Nalgonda","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nirmal","city":"Nirmal","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Nizamabad","city":"Nizamabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Peddapally","city":"Ramagundam","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Rangareddy","city":"Shadnagar","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Sangareddy","city":"Sangareddy","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Siddipet","city":"Siddipet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Suryapet","city":"Suryapet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Vikarabad","city":"Vikarabad","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Wanapathy","city":"Wanaparthy","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Warangal","city":"Narsampet","climate":"Hot and Dry"},
    {"state":"Telangana","district":"Yadadri Bhuvanagiri","city":"Bhongir","climate":"Hot and Dry"},
    {"state":"Tripura","district":"Dhalai","city":"Ambassa","climate":"Warm & humid"},
    {"state":"Tripura","district":"Gomati","city":"Udaipur","climate":"Warm & humid"},
    {"state":"Tripura","district":"North Tripura","city":"Dharmanagar","climate":"Warm & humid"},
    {"state":"Tripura","district":"South Tripura","city":"Belonia","climate":"Warm & humid"},
    {"state":"Tripura","district":"West Tripura","city":"Agartala","climate":"Warm & humid"},
    {"state":"Uttar Pradesh","district":"Agra","city":"Agra","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Aligarh","city":"Aligarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ambedkar Nagar","city":"Akbarpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Amroha","city":"Amroha","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Auraiya","city":"Auraiya","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ayodhya","city":"Ayodhya","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Azamgarh","city":"Azamgarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Baghpat","city":"Baghpat","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bahraich","city":"Bahraich","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ballia","city":"Ballia","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Balrampur","city":"Balrampur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Banda","city":"Banda","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Barabanki","city":"Barabanki","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bareilly","city":"Bareilly","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Basti","city":"Basti","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bhadohi","city":"Bhadohi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bijnor","city":"Bijnor","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Budaun","city":"Budaun","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Bulandshahar","city":"Bulandshahr","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Chandauli","city":"Chandauli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Chitrakoot","city":"Chitrakoot Dham","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Deoria","city":"Deoria","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Etah","city":"Etah","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Etawah","city":"Etawah","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Farrukhabad","city":"Farrukhabad-Cum-Fatehgarh","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Fatehpur","city":"Fatehpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Firozabad","city":"Firozabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gautam Buddha Nagar","city":"Noida","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ghaziabad","city":"Ghaziabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Ghazipur","city":"Ghazipur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gonda","city":"Gonda","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Gorakhpur","city":"Gorakhpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hamirpur","city":"Hamirpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hapur","city":"Hapur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hardoi","city":"Hardoi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Hathras","city":"Hathras","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jalaun","city":"Jalaun","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jaunpur","city":"Jaunpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Jhansi","city":"Jhansi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kannauj","city":"Kannauj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kanpur","city":"Kanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kanpur Dehat","city":"Akbarpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kasganj","city":"Kasganj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kaushambi","city":"Kaushambi","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kheeri","city":"Lakhimpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Kushinagar","city":"Kushinagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Lalitpur","city":"Lalitpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Lucknow","city":"Lucknow","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Maharajganj","city":"Maharajganj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mahoba","city":"Mahoba","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mainpuri","city":"Mainpuri","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mathura","city":"Mathura-Vrindavan","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mau","city":"Maunath Bhanjan","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Meerut","city":"Meerut","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Mirzapur-Cum-Vindhyachal","city":"Mirzapur-Cum-Vindhyachal","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Moradabad","city":"Moradabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Muzaffarnagar","city":"Muzaffarnagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Pilibhit","city":"Pilibhit","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Pratapgarh","city":"Pratapgarh City","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Prayagraj","city":"Prayagraj","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Raebareli","city":"Rae Bareli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Rampur","city":"Rampur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Saharanpur","city":"Saharanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sambhal","city":"Sambhal","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sant Kabir Nagar","city":"Khalilabad","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Shahjahanpur","city":"Shahjahanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Shamli","city":"Shamli","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Siddharthnagar","city":"Siddharthnagar","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sitapur","city":"Sitapur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sonbhadra","city":"Sonbhadra","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Sultanpur","city":"Sultanpur","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Unnao","city":"Unnao","climate":"Composite"},
    {"state":"Uttar Pradesh","district":"Varanasi","city":"Varanasi","climate":"Composite"},
    {"state":"Uttarakhand","district":"Almora","city":"Almora","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Bageshwar","city":"Bageshwar","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Chamoli Gopeshwar","city":"Chamoli-Gopeshwar","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Champawat","city":"Champawat","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Dehradun","city":"Dehradun","climate":"Composite"},
    {"state":"Uttarakhand","district":"Hardwar","city":"Hardwar","climate":"Composite"},
    {"state":"Uttarakhand","district":"Nainital","city":"Haldwani","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Nainital","city":"Nainital","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Pauri Garhwal","city":"Kotdwara","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Pithoragarh","city":"Pithoragarh","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Rudraprayag","city":"Rudraprayag","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Tehri","city":"Tehri","climate":"Temperate"},
    {"state":"Uttarakhand","district":"Udhamsingh Nagar","city":"Rudrapur","climate":"Composite"},
    {"state":"Uttarakhand","district":"Uttarkashi","city":"Uttarkashi","climate":"Temperate"},
    {"state":"West Bengal","district":"Bankura","city":"Bankura","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Barddhamam","city":"Asansol","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Barddhamam","city":"Barddhaman","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Birbhum","city":"Suri","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Dakshin Dinajpur","city":"Balurghat","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Darjeeling","city":"Darjiling","climate":"Cold"},
    {"state":"West Bengal","district":"Darjeeling","city":"Siliguri","climate":"Cold"},
    {"state":"West Bengal","district":"Hooghly","city":"Chandannagar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Hooghly","city":"Serampore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Howrah","city":"Haora","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Jalpaiguri","city":"Jalpaiguri","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Koch Bihar","city":"Koch Bihar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Kolkata","city":"Kolkata","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Maldah","city":"English Bazar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Murshidabad","city":"Berhampore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Nadia","city":"Krishnanagar","climate":"Warm & humid"},
    {"state":"West Bengal","district":"North 24 Parganas","city":"Barasat","climate":"Warm & humid"},
    {"state":"West Bengal","district":"North 24 Parganas","city":"Barrackpore","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Paschim Medinipur","city":"Medinipur","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Purba Midnapur","city":"Haldia","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Puruliya","city":"Puruliya","climate":"Warm & humid"},
    {"state":"West Bengal","district":"South 24 Parganas","city":"Diamond Harbour","climate":"Warm & humid"},
    {"state":"West Bengal","district":"Uttar Dinajpur","city":"Raiganj","climate":"Warm & humid"},
    {"state":"Lakshadweep","district":"Lakshadweep","city":"Lakshadweep","climate":"Warm & humid"},
]



# ─── HELPERS ──────────────────────────────────────────────────────────────────
def fuel_to_co2e(fuel_name, quantity, unit):
    if quantity <= 0 or fuel_name not in FUEL_EF:
        return 0.0
    ef = FUEL_EF[fuel_name]
    co2_factor, ch4_factor, n2o_factor, ncv, density = ef
    if fuel_name == "Electricity":
        return quantity * co2_factor
    if fuel_name == "MSW incineration":
        return quantity * (co2_factor + ch4_factor * GWP["CH4"] + n2o_factor * GWP["N2O"])
    if unit == "MWh":
        tj = quantity * 0.0036
    elif unit == "kL":
        if density is None: return 0.0
        tj = quantity * density * ncv
    elif unit == "tonne":
        tj = quantity * ncv
    elif unit == "TJ":
        tj = quantity
    else:
        tj = 0.0
    return tj * (co2_factor + ch4_factor * GWP["CH4"] + n2o_factor * GWP["N2O"])

def project_bau(base, r, n):
    return {s: max(v, 0) * (1 + r) ** n for s, v in base.items()}

def apply_mitigation(bau, strat):
    return {s: bau[s] * (1 - strat.get(s, 0)) for s in bau}

def timeseries(base, r, by, yrs, ep, ha):
    rows = []
    for yr in yrs:
        n = yr - by
        b = project_bau(base, r, n)
        e = apply_mitigation(b, ep)
        h = apply_mitigation(b, ha)
        rows.append({
            "Year": yr,
            "Reference": sum(b.values()),
            "Existing & Planned": sum(e.values()),
            "High Ambition": sum(h.values()),
            **{f"BAU_{s}": v for s, v in b.items()},
            **{f"HA_{s}": v for s, v in h.items()},
        })
    return pd.DataFrame(rows)

def budget_table(base, ha, r, n):
    bau = project_bau(base, r, n)
    rows = []
    for s, frac in ha.items():
        red = bau[s] * frac
        rows.append({
            "Sector": s,
            "BAU (t CO2e)": round(bau[s]),
            "Reduction %": f"{frac*100:.0f}%",
            "GHG Reduced (t CO2e)": round(red),
            "Investment (Crore)": round(red / 1e6 * UNIT_COST.get(s, 10.0), 1)
        })
    df = pd.DataFrame(rows)
    total = pd.DataFrame([{
        "Sector": "TOTAL",
        "BAU (t CO2e)": round(df["BAU (t CO2e)"].sum()),
        "Reduction %": "",
        "GHG Reduced (t CO2e)": round(df["GHG Reduced (t CO2e)"].sum()),
        "Investment (Crore)": round(df["Investment (Crore)"].sum(), 1)
    }])
    return pd.concat([df, total], ignore_index=True)

def compute_emissions(d):
    """Compute all sector emissions from form data dict."""
    b = {}
    # Buildings
    for prefix, sector in [("res","Buildings – Residential"),("com","Buildings – Commercial"),
                            ("ins","Buildings – Public & Inst."),("ind","Buildings – Industrial")]:
        fuels = [("Electricity","MWh"),("Firewood","tonne"),("Kerosene","kL"),
                 ("PNG","tonne"),("LPG","tonne")]
        if prefix == "ind":
            fuels.append(("Coal","tonne"))
        b[sector] = sum(fuel_to_co2e(f, float(d.get(f"{prefix}_{f.replace(' ','_')}",0)), u)
                        for f,u in fuels)
    # Electricity Generation
    b["Electricity Generation"] = (
        fuel_to_co2e("Natural gas (TJ)", float(d.get("ng_tj",0)), "TJ") +
        fuel_to_co2e("Coal", float(d.get("coal_tj",0)), "TJ") +
        fuel_to_co2e("MSW incineration", float(d.get("msw_pw",0)), "MWh")
    )
    # Transport
    b["Transport – Road"] = (
        fuel_to_co2e("Petrol", float(d.get("t_pet",0)), "kL") +
        fuel_to_co2e("Diesel", float(d.get("t_die",0)), "kL") +
        fuel_to_co2e("CNG", float(d.get("t_cng",0)), "tonne") +
        fuel_to_co2e("Auto LPG", float(d.get("t_alpg",0)), "tonne") +
        fuel_to_co2e("Electricity", float(d.get("t_elec",0)), "MWh")
    )
    b["Transport – Rail"] = (
        fuel_to_co2e("Diesel", float(d.get("r_die",0)), "kL") +
        fuel_to_co2e("Electricity", float(d.get("r_elec",0)), "MWh")
    )
    b["Transport – Water/Aviation"] = (
        fuel_to_co2e("Petrol", float(d.get("w_pet",0)), "kL") +
        fuel_to_co2e("Diesel", float(d.get("w_die",0)), "kL") +
        fuel_to_co2e("Aviation gasoline", float(d.get("av_gas",0)), "kL") +
        fuel_to_co2e("Jet kerosene", float(d.get("av_jet",0)), "kL")
    )
    # Solid Waste
    sw_total = float(d.get("sw_tot",0))
    sw_annual = sw_total * 365
    sw_lf_m = float(d.get("sw_lfm",50))
    sw_lf_u = float(d.get("sw_lfu",20))
    sw_inc  = float(d.get("sw_inc",5))
    sw_com  = float(d.get("sw_com",10))
    DOC_f, MCF_m, MCF_u, OX, F_CH4 = 0.5, 1.0, 0.6, 0.1, 0.5
    doc_m = sw_annual * sw_lf_m/100 * DOC_f * MCF_m
    doc_u = sw_annual * sw_lf_u/100 * DOC_f * MCF_u
    ch4_lf = (doc_m + doc_u) * F_CH4 * (16/12) * (1 - OX)
    n2o_inc = sw_annual * sw_inc/100 * 60e-6 * GWP["N2O"]
    ch4_bio = sw_annual * float(d.get("sw_com",10))/100 * 4e-6 * GWP["CH4"]
    b["Waste – Solid Waste"] = ch4_lf * GWP["CH4"] + n2o_inc
    b["Waste – Biological"]  = ch4_bio
    # Wastewater
    population = float(d.get("population",500000))
    ww_bod  = float(d.get("ww_bod",34))
    ww_prot = float(d.get("ww_prot",21.54))
    BOD_total = population * ww_bod / 1000 * 365
    EF_j = {"aerobic":0.1,"uasb":0.3,"septic":0.5,"open":0.8}
    ch4_ww = BOD_total * (
        float(d.get("ww_aer",30))/100 * EF_j["aerobic"] +
        float(d.get("ww_uasb",20))/100 * EF_j["uasb"] +
        float(d.get("ww_sep",20))/100 * EF_j["septic"] +
        float(d.get("ww_open",30))/100 * EF_j["open"]
    )
    N_effluent = population * ww_prot * 0.16 * 1.1
    n2o_ww = N_effluent * 0.005 * 44/28
    b["Waste – Wastewater"] = ch4_ww * GWP["CH4"] + n2o_ww * GWP["N2O"]
    # AFOLU
    af_dairy=float(d.get("af_dc",0)); af_nond=float(d.get("af_ndc",0))
    af_buf=float(d.get("af_bufd",0)); af_bufnd=float(d.get("af_bufnd",0))
    af_sheep=float(d.get("af_sheep",0)); af_goat=float(d.get("af_goat",0))
    af_swine=float(d.get("af_swine",0))
    ch4_ent = (af_dairy*28 + af_nond*23 + af_buf*43 + af_bufnd*33 +
               af_sheep*5 + af_goat*5 + af_swine*1.5)
    af_forest_d=float(d.get("af_fd",0)); af_forest_m=float(d.get("af_fm",0))
    af_forest_o=float(d.get("af_fo",0)); af_wet=float(d.get("af_wet",0))
    C_seq = (af_forest_d*6 + af_forest_m*4 + af_forest_o*2) * 0.47 * 44/12
    ch4_rice = af_wet * 1.3 * 120 / 1000
    b["AFOLU"] = (ch4_ent/1000 * GWP["CH4"]) + (ch4_rice * GWP["CH4"]) - C_seq
    # IPPU
    ip_clink = float(d.get("ip_clink",0))
    ip_cfrac_str = d.get("ip_cfrac","OPC (0.95)")
    ip_cfrac_val = float(ip_cfrac_str.split("(")[1].replace(")",""))
    b["IPPU"] = (
        ip_clink * ip_cfrac_val * 0.507 +
        float(d.get("ip_lime",0)) * 0.754 +
        float(d.get("ip_glass",0)) * 0.2 * (1 - float(d.get("ip_cullet",0.2))) +
        float(d.get("ip_ls",0)) * 0.44 +
        float(d.get("ip_nh3",0)) * 1.694 +
        float(d.get("ip_hno3",0)) * 9.0 * GWP["N2O"] / 1000 +
        float(d.get("ip_soda",0)) * 0.138 +
        float(d.get("ip_bof",0)) * 0.26 +
        float(d.get("ip_eaf",0)) * 0.14 +
        float(d.get("ip_alpb",0)) * 1.6 +
        float(d.get("ip_also",0)) * 2.2 +
        float(d.get("ip_hfc",0)) * 0.05 * 2000 +
        float(d.get("ip_sf6",0)) * 0.02 * 23500
    )
    return {s: max(0.0, v) if s != "AFOLU" else v for s, v in b.items()}

CT = dict(
    template="plotly_white",
    font=dict(family="Inter, sans-serif", size=12, color="#1A1A2E"),
    paper_bgcolor="#FFFFFF",
    plot_bgcolor="#FAFBFC",
    title_font=dict(family="Montserrat, sans-serif", size=14, color="#1B2A4A"),
)

def make_charts(df_ts, base_emissions, bdf, city_label, target_pct, base_year):
    charts = {}
    # Trajectory
    fig = go.Figure()
    cfgs = {
        "Reference":          ("#DC2626","solid",3.0),
        "Existing & Planned": ("#D97706","dash",2.2),
        "High Ambition":      ("#059669","longdash",2.5),
    }
    for name,(col,dash,w) in cfgs.items():
        fig.add_trace(go.Scatter(
            x=df_ts["Year"], y=(df_ts[name]/1e6).round(4), name=name,
            mode="lines+markers",
            line=dict(color=col,dash=dash,width=w),
            marker=dict(size=7,color=col,line=dict(width=1.5,color="#fff"))
        ))
    if target_pct > 0:
        brow = df_ts[df_ts["Year"]==base_year]["Reference"].values
        if len(brow):
            tval = brow[0]*(1-target_pct/100)/1e6
            fig.add_hline(y=tval,line_dash="dot",line_color="#7C3AED",line_width=2,
                          annotation_text=f" Net-zero target ({target_pct:.0f}% reduction)",
                          annotation_font_color="#7C3AED",annotation_font_size=11)
    fig.update_layout(title=f"<b>GHG Emission Scenarios — {city_label}</b>",
                      xaxis_title="Year",yaxis_title="GHG Emissions (Mt CO₂e)",
                      hovermode="x unified",height=420,
                      legend=dict(orientation="h",yanchor="bottom",y=1.02,xanchor="right",x=1),
                      **CT)
    charts["trajectory"] = json.loads(plotly.utils.PlotlyJSONEncoder().encode(fig))

    # Pie
    pos = {s:v for s,v in base_emissions.items() if v>0}
    if pos:
        fig2 = go.Figure(go.Pie(
            labels=list(pos.keys()),
            values=[round(v/1e3,2) for v in pos.values()],
            marker=dict(colors=[SECTOR_COLORS[s] for s in pos],
                        line=dict(color="#FFFFFF",width=2)),
            hole=0.44,textinfo="label+percent",
            hovertemplate="<b>%{label}</b><br>%{value:.1f} kt CO₂e<extra></extra>"
        ))
        fig2.update_layout(title="<b>Base Year Emission Profile</b>",height=400,**CT)
        charts["pie"] = json.loads(plotly.utils.PlotlyJSONEncoder().encode(fig2))

    # Bar by group
    grp_map = {
        "Buildings":   ["Buildings – Residential","Buildings – Commercial",
                        "Buildings – Public & Inst.","Buildings – Industrial"],
        "Electricity": ["Electricity Generation"],
        "Transport":   ["Transport – Road","Transport – Rail","Transport – Water/Aviation"],
        "Waste":       ["Waste – Solid Waste","Waste – Biological","Waste – Wastewater"],
        "AFOLU":       ["AFOLU"],
        "IPPU":        ["IPPU"],
    }
    grps,vals,cols=[],[],[]
    for grp,slist in grp_map.items():
        tot = sum(base_emissions.get(s,0) for s in slist)/1e6
        grps.append(grp); vals.append(round(tot,4))
        cols.append(SECTOR_COLORS.get(slist[0],"#888"))
    fig3 = go.Figure(go.Bar(x=grps,y=vals,marker_color=cols,
        text=[f"{v:.4f}" for v in vals],textposition="outside",
        marker_line=dict(width=0)))
    fig3.update_layout(title="<b>Emissions by Sector Group (Mt CO₂e)</b>",
                       yaxis_title="Mt CO₂e",height=360,**CT)
    charts["bar_group"] = json.loads(plotly.utils.PlotlyJSONEncoder().encode(fig3))

    # Budget chart
    df_b = bdf[bdf["Sector"]!="TOTAL"]
    fig4 = make_subplots(specs=[[{"secondary_y":True}]])
    fig4.add_trace(go.Bar(
        x=df_b["Sector"],y=df_b["Investment (Crore)"],name="Investment (₹ Cr)",
        marker_color=[SECTOR_COLORS.get(s,"#888") for s in df_b["Sector"]],
        marker_line=dict(width=0),opacity=0.9,
    ),secondary_y=False)
    fig4.add_trace(go.Scatter(
        x=df_b["Sector"],y=df_b["GHG Reduced (t CO2e)"]/1e6,
        name="GHG Reduced (Mt)",mode="lines+markers",
        line=dict(color="#1B2A4A",width=2.5),
        marker=dict(size=9,color="#1B2A4A",line=dict(width=2,color="#fff")),
    ),secondary_y=True)
    fig4.update_layout(title="<b>Investment vs GHG Reduction</b>",height=380,**CT)
    charts["budget"] = json.loads(plotly.utils.PlotlyJSONEncoder().encode(fig4))

    return charts


# ─── ROUTES ───────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    states = sorted(set(c["state"] for c in INDIA_CITIES))
    return render_template("index.html", states=states, cities=INDIA_CITIES, sectors=SECTORS)

@app.route("/api/cities/<state>")
def get_cities(state):
    cities = [c for c in INDIA_CITIES if c["state"] == state]
    return jsonify(cities)

@app.route("/api/calculate", methods=["POST"])
def calculate():
    d = request.json
    base_emissions = compute_emissions(d)
    base_year   = int(d.get("base_year", 2025))
    target_year = int(d.get("target_year", 2050))
    interim1    = int(d.get("interim1", 2030))
    interim2    = int(d.get("interim2", 2040))
    growth_rate = float(d.get("growth_rate", 2.0)) / 100
    target_pct  = float(d.get("target_pct", 60))
    population  = float(d.get("population", 500000))
    area_sqkm   = float(d.get("area_sqkm", 150))
    city_label  = f"{d.get('district','City')}, {d.get('state','State')}"

    ep_strategies = {s: float(d.get(f"ep_{s}", 8)) / 100 for s in SECTORS}
    ha_strategies = {s: float(d.get(f"ha_{s}", 25)) / 100 for s in SECTORS}

    years_list = sorted(set([base_year,interim1,interim2,target_year] +
                             list(range(base_year, target_year+1))))
    df_ts  = timeseries(base_emissions, growth_rate, base_year,
                        years_list, ep_strategies, ha_strategies)
    n_years = target_year - base_year
    bdf    = budget_table(base_emissions, ha_strategies, growth_rate, n_years)

    base_total = sum(base_emissions.values())
    bau_end    = df_ts["Reference"].iloc[-1]
    ha_end     = df_ts["High Ambition"].iloc[-1]
    ha_saving  = bau_end - ha_end
    total_inv  = bdf[bdf["Sector"]=="TOTAL"]["Investment (Crore)"].values[0]

    charts = make_charts(df_ts, base_emissions, bdf, city_label, target_pct, base_year)

    # Milestone table
    milestones = []
    for yr in [interim1, interim2, target_year]:
        row = df_ts[df_ts["Year"]==yr]
        if not row.empty:
            bau_yr = row["Reference"].values[0]
            ha_yr  = row["High Ambition"].values[0]
            pct_req = target_pct * (yr-base_year) / n_years if n_years else 0
            ach    = (1 - ha_yr/bau_yr)*100 if bau_yr!=0 else 0
            milestones.append({
                "year": yr,
                "bau": round(bau_yr/1e6,4),
                "target": round(bau_yr*(1-pct_req/100)/1e6,4),
                "ha": round(ha_yr/1e6,4),
                "required_pct": round(pct_req,1),
                "achieved_pct": round(ach,1),
                "status": "On Track" if ach>=pct_req else "Gap"
            })

    sector_detail = [
        {"sector": s, "emissions": round(v), "share": round(v/base_total*100,2) if base_total else 0}
        for s,v in base_emissions.items()
    ]
    budget_rows = bdf.to_dict("records")

    return jsonify({
        "kpis": {
            "base_total_mt": round(base_total/1e6,4),
            "per_capita": round(base_total/population,2) if population else 0,
            "per_sqkm": round(base_total/area_sqkm/1e3,2) if area_sqkm else 0,
            "bau_end_mt": round(bau_end/1e6,4),
            "ha_end_mt": round(ha_end/1e6,4),
            "total_inv": round(total_inv,0),
            "ha_saving_pct": round(ha_saving/bau_end*100,1) if bau_end else 0,
            "target_year": target_year,
        },
        "charts": charts,
        "milestones": milestones,
        "sector_detail": sector_detail,
        "budget": budget_rows,
        "city_label": city_label,
    })

@app.route("/api/download/csv", methods=["POST"])
def download_csv():
    d = request.json
    base_emissions = compute_emissions(d)
    base_year   = int(d.get("base_year",2025))
    target_year = int(d.get("target_year",2050))
    interim1    = int(d.get("interim1",2030))
    interim2    = int(d.get("interim2",2040))
    growth_rate = float(d.get("growth_rate",2.0))/100
    ep_strategies = {s: float(d.get(f"ep_{s}",8))/100 for s in SECTORS}
    ha_strategies = {s: float(d.get(f"ha_{s}",25))/100 for s in SECTORS}
    years_list = sorted(set([base_year,interim1,interim2,target_year] +
                             list(range(base_year,target_year+1))))
    df_ts = timeseries(base_emissions,growth_rate,base_year,years_list,ep_strategies,ha_strategies)
    exp = df_ts[["Year"]].copy()
    exp["Reference (Mt)"]          = (df_ts["Reference"]/1e6).round(4)
    exp["Existing & Planned (Mt)"] = (df_ts["Existing & Planned"]/1e6).round(4)
    exp["High Ambition (Mt)"]      = (df_ts["High Ambition"]/1e6).round(4)
    buf = io.StringIO()
    exp.to_csv(buf, index=False)
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode()),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"ascent_{d.get('district','city')}_scenarios.csv"
    )

@app.route("/api/template")
def download_template():
    template = {
        "city_name": "Your City", "state": "Gujarat", "district": "Surat",
        "governance_tier": "City / ULB", "climate_zone": "Hot and Dry",
        "population": 500000, "area_sqkm": 150,
        "base_year": 2025, "interim_year_1": 2030,
        "interim_year_2": 2040, "target_year": 2050,
        "population_growth_rate_pct": 2.0,
        "emissions_tCO2e": {s: 0 for s in SECTORS},
        "ep_strategies_pct": {s: 8 for s in SECTORS},
        "ha_strategies_pct": {s: 25 for s in SECTORS},
    }
    buf = io.BytesIO(json.dumps(template, indent=2).encode())
    buf.seek(0)
    return send_file(buf, mimetype="application/json",
                     as_attachment=True, download_name="ascent_template.json")


@app.route("/api/download/excel", methods=["POST"])
def download_excel():
    """Download scenario data as Excel workbook with multiple sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    d = request.json
    base_emissions = compute_emissions(d)
    base_year   = int(d.get("base_year", 2025))
    target_year = int(d.get("target_year", 2050))
    interim1    = int(d.get("interim1", 2030))
    interim2    = int(d.get("interim2", 2040))
    growth_rate = float(d.get("growth_rate", 2.0)) / 100
    target_pct  = float(d.get("target_pct", 60))
    ep_strategies = {s: float(d.get(f"ep_{s}", 8)) / 100 for s in SECTORS}
    ha_strategies = {s: float(d.get(f"ha_{s}", 25)) / 100 for s in SECTORS}
    years_list = sorted(set([base_year, interim1, interim2, target_year] + list(range(base_year, target_year+1))))
    df_ts = timeseries(base_emissions, growth_rate, base_year, years_list, ep_strategies, ha_strategies)
    n_years = target_year - base_year
    bdf = budget_table(base_emissions, ha_strategies, growth_rate, n_years)

    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FILL = PatternFill("solid", fgColor="1B2A4A")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    TEAL_FILL = PatternFill("solid", fgColor="4AADA8")
    TOTAL_FILL = PatternFill("solid", fgColor="E6F4EA")
    BOLD = Font(bold=True)

    def style_headers(ws, row, ncols, fill=None, font=None):
        fill = fill or HDR_FILL
        font = font or HDR_FONT
        for c in range(1, ncols+1):
            cell = ws.cell(row, c)
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    def auto_width(ws):
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 42)

    # Sheet 1: Scenarios
    ws1 = wb.active; ws1.title = "Scenarios"
    h1 = ["Year","Reference (Mt CO2e)","Existing & Planned (Mt)","High Ambition (Mt)"]
    for ci,h in enumerate(h1,1): ws1.cell(1,ci,h)
    style_headers(ws1,1,len(h1))
    for ri,(_,r) in enumerate(df_ts.iterrows(),2):
        vals = [int(r["Year"]), round(r["Reference"]/1e6,4), round(r["Existing & Planned"]/1e6,4), round(r["High Ambition"]/1e6,4)]
        for ci,v in enumerate(vals,1):
            ws1.cell(ri,ci,v).border=BORDER
    auto_width(ws1)

    # Sheet 2: Base Emissions
    ws2 = wb.create_sheet("Base Emissions")
    h2 = ["Sector","Emissions (t CO2e)","Share (%)"]
    for ci,h in enumerate(h2,1): ws2.cell(1,ci,h)
    style_headers(ws2,1,len(h2))
    total_base = sum(base_emissions.values())
    for ri,(s,v) in enumerate(base_emissions.items(),2):
        share = round(v/total_base*100,2) if total_base else 0
        for ci,val in enumerate([s,round(v),share],1):
            ws2.cell(ri,ci,val).border=BORDER
    tr = len(base_emissions)+2
    for ci,val in enumerate(["TOTAL",round(total_base),100.0],1):
        c = ws2.cell(tr,ci,val); c.fill=TOTAL_FILL; c.font=BOLD; c.border=BORDER
    auto_width(ws2)

    # Sheet 3: Mitigation Budget
    ws3 = wb.create_sheet("Mitigation Budget")
    h3 = ["Sector","BAU (t CO2e)","Reduction %","GHG Reduced (t CO2e)","Investment (Cr INR)"]
    for ci,h in enumerate(h3,1): ws3.cell(1,ci,h)
    style_headers(ws3,1,len(h3),TEAL_FILL)
    for ri,row in enumerate(bdf.to_dict("records"),2):
        vals = [row["Sector"],row["BAU (t CO2e)"],row["Reduction %"],row["GHG Reduced (t CO2e)"],row["Investment (Crore)"]]
        for ci,v in enumerate(vals,1):
            c = ws3.cell(ri,ci,v); c.border=BORDER
            if row["Sector"]=="TOTAL": c.fill=TOTAL_FILL; c.font=BOLD
    auto_width(ws3)

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    city_label = f"{d.get('district','City')}, {d.get('state','State')}"
    rows = [
        ("City / District", city_label),
        ("Population", d.get("population","")),
        ("Area (km²)", d.get("area_sqkm","")),
        ("Base Year", base_year), ("Target Year", target_year),
        ("Growth Rate (%/yr)", d.get("growth_rate","")),
        ("Target Reduction (%)", target_pct),
        ("Base Emissions (Mt CO₂e)", round(total_base/1e6,4)),
        ("BAU at Target Year (Mt)", round(df_ts["Reference"].iloc[-1]/1e6,4)),
        ("High Ambition (Mt)", round(df_ts["High Ambition"].iloc[-1]/1e6,4)),
        ("Total HA Investment (₹ Cr)", round(float(bdf[bdf["Sector"]=="TOTAL"]["Investment (Crore)"].values[0]),1)),
    ]
    ws4.column_dimensions["A"].width=40; ws4.column_dimensions["B"].width=28
    for ri,(k,v) in enumerate(rows,1):
        ws4.cell(ri,1,k).font=BOLD; ws4.cell(ri,2,v)
        ws4.cell(ri,1).border=BORDER; ws4.cell(ri,2).border=BORDER

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    city_safe = d.get("district","city").replace(" ","_").replace("/","_")
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"ASCENT_{city_safe}_{base_year}_{target_year}.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
